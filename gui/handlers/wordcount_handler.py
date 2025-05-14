"""
提供字数统计相关的业务逻辑处理
"""
import os
import threading
from datetime import datetime
from docx import Document
from heading_utils import get_document_stats
from word_processors import extract_author_number, extract_author_from_filename

# 检查是否安装了openpyxl库
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class WordcountHandler:
    """字数统计处理器，处理字数检测和Excel报告相关功能"""
    
    def __init__(self, app):
        """
        初始化字数统计处理器
        
        参数:
            app: 主应用程序实例
        """
        self.app = app
        
    def validate_paths(self, input_dir, output_dir=None):
        """
        验证路径是否有效
        
        参数:
            input_dir: 输入目录
            output_dir: 输出目录(可选)
            
        返回:
            tuple: (是否有效, 错误信息)
        """
        if not input_dir:
            return False, "请选择输入文件夹"
            
        # 检查输入目录是否存在
        if not os.path.exists(input_dir):
            return False, "输入目录不存在"
        
        # 如果提供了输出目录，则检查
        if output_dir is not None and not output_dir:
            return False, "请选择输出文件夹"
            
        return True, ""
        
    def check_wordcount(self):
        """开始字数检测"""
        # 获取路径配置
        paths = self.app.file_frame.get_paths()
        input_dir = paths["input_dir"]
        output_dir = paths["output_dir"]
        
        # 验证路径
        valid, error_msg = self.validate_paths(input_dir, output_dir)
        if not valid:
            self.app.set_status(error_msg)
            return
        
        # 获取字数检测配置
        wordcount_config = self.app.wordcount_frame.get_wordcount_config()
        min_words = wordcount_config["min_words"]
        export_excel = wordcount_config["export_excel"]
        
        # 准备开始字数检测
        self.app.reset_for_processing()
        self.app.set_status("正在检测字数...")
        
        # 在新线程中运行字数检测
        thread = threading.Thread(
            target=self._check_thread, 
            args=(input_dir, output_dir, min_words, export_excel),
            daemon=True
        )
        thread.start()
    
    def _check_thread(self, input_dir, output_dir, min_words, export_excel):
        """字数检测线程"""
        try:
            # 获取所有Word文件
            docx_files = [f for f in os.listdir(input_dir) if f.endswith(('.doc', '.docx'))]
            word_counts = []
            low_wordcount_files = []
            
            total_files = len(docx_files)
            
            # 配置进度条最大值
            self.app.root.after(0, lambda: self.app.progress_bar.configure(maximum=total_files))
            
            for index, filename in enumerate(docx_files, 1):
                # 更新进度条
                current_progress = index / total_files * 100
                self.app.root.after(0, lambda p=index: self.app.progress_bar.configure(value=p))
                self.app.root.after(0, lambda: self.app.progress_bar.update())
                
                self.app.set_status(f"正在检测文件 {index}/{total_files} ({int(current_progress)}%)...")
                
                input_file = os.path.join(input_dir, filename)
                try:
                    doc = Document(input_file)
                    stats = get_document_stats(doc)
                    word_count = stats['word_count']
                    para_count = stats['paragraph_count']
                    char_count = stats['character_count']
                    
                    # 提取作者名
                    author_name = extract_author_from_filename(filename) or "未知"
                    
                    word_counts.append((filename, word_count, para_count, char_count, author_name))
                    if word_count < min_words:
                        low_wordcount_files.append((filename, word_count, author_name))
                except Exception as e:
                    self.app.log_text.insert('end', f"× {filename}: 字数检测失败 - {str(e)}\n")
            
            # 按字数排序
            word_counts.sort(key=lambda x: x[1])
            
            # 显示结果
            self._show_results(word_counts, low_wordcount_files, min_words)
            
            # 导出Excel报告
            excel_path = ""
            if export_excel and word_counts:
                try:
                    excel_path = self._export_wordcount_excel(word_counts, low_wordcount_files, min_words, output_dir)
                    self.app.log_text.insert('end', f"\n\nExcel报告已导出至: {excel_path}\n")
                except Exception as e:
                    self.app.log_text.insert('end', f"\n\nExcel报告导出失败: {str(e)}\n")
            
            self.app.log_text.see('end')
            
            # 设置进度条为100%完成
            self.app.root.after(0, lambda: self.app.progress_bar.configure(value=total_files))
            
            # 更新状态
            self.app.root.after(0, lambda: self.app.set_status(
                f"字数检测完成，共 {len(word_counts)} 个文件，其中 {len(low_wordcount_files)} 个不足 {min_words} 字" +
                (f"，已生成Excel报告" if excel_path else "")
            ))
            self.app.root.after(0, self.app.enable_buttons)
            
        except Exception as e:
            self.app.root.after(0, lambda: self.app.set_status(f"字数检测出错: {str(e)}"))
            self.app.root.after(0, self.app.enable_buttons)
    
    def _show_results(self, word_counts, low_wordcount_files, min_words):
        """显示字数检测结果"""
        self.app.log_text.insert('end', f"字数统计报告 (最小字数要求: {min_words})\n")
        self.app.log_text.insert('end', "="*50 + "\n\n")
        
        # 显示所有文件的字数
        self.app.log_text.insert('end', "所有文件的字数统计:\n")
        for filename, word_count, para_count, char_count, author_name in word_counts:
            status = "不足" if word_count < min_words else "合格"
            self.app.log_text.insert('end', f"{filename} ({author_name}): {word_count} 字 ({para_count} 段落, {char_count} 字符) - {status}\n")
        
        # 显示字数不足的文件
        if low_wordcount_files:
            self.app.log_text.insert('end', "\n\n字数不足的文件:\n")
            for filename, word_count, author_name in low_wordcount_files:
                self.app.log_text.insert('end', f"{filename} ({author_name}): {word_count} 字\n")
        
        # 统计信息
        summary = f"\n\n统计信息:\n"
        summary += f"总文件数: {len(word_counts)} 个\n"
        summary += f"平均字数: {sum(x[1] for x in word_counts) / len(word_counts) if word_counts else 0:.1f} 字\n"
        summary += f"最大字数: {max(x[1] for x in word_counts) if word_counts else 0} 字\n"
        summary += f"最小字数: {min(x[1] for x in word_counts) if word_counts else 0} 字\n"
        summary += f"字数不足文件: {len(low_wordcount_files)} 个\n"
        summary += f"字数合格文件: {len(word_counts) - len(low_wordcount_files)} 个\n"
        
        self.app.log_text.insert('end', summary)
    
    def _export_wordcount_excel(self, word_counts, low_wordcount_files, min_words, output_dir):
        """导出字数统计Excel报告"""
        if not EXCEL_AVAILABLE:
            raise ImportError("缺少openpyxl库，无法导出Excel")
        
        # 创建报告目录
        reports_dir = os.path.join(output_dir, "字数统计报告")
        os.makedirs(reports_dir, exist_ok=True)
        
        # 创建带有当前日期时间的文件名
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(reports_dir, f"字数统计报告_{current_time}.xlsx")
        
        # 创建新的Excel工作簿
        wb = openpyxl.Workbook()
        
        # 创建总览工作表
        overview_sheet = wb.active
        overview_sheet.title = "字数统计总览"
        
        # 设置标题行
        headers = ["文件名", "作者", "字数", "段落数", "字符数", "状态"]
        for col, header in enumerate(headers, 1):
            cell = overview_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row, (filename, word_count, para_count, char_count, author_name) in enumerate(word_counts, 2):
            status = "不足" if word_count < min_words else "合格"
            
            overview_sheet.cell(row=row, column=1).value = filename
            overview_sheet.cell(row=row, column=2).value = author_name
            overview_sheet.cell(row=row, column=3).value = word_count
            overview_sheet.cell(row=row, column=4).value = para_count
            overview_sheet.cell(row=row, column=5).value = char_count
            overview_sheet.cell(row=row, column=6).value = status
            
            # 设置字数不足的行底色为浅红色
            if word_count < min_words:
                for col in range(1, 7):
                    overview_sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
        
        # 创建字数不足工作表
        if low_wordcount_files:
            low_sheet = wb.create_sheet(title="字数不足文件")
            
            # 设置标题行
            headers = ["文件名", "作者", "字数", "差额"]
            for col, header in enumerate(headers, 1):
                cell = low_sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 写入数据
            for row, (filename, word_count, author_name) in enumerate(low_wordcount_files, 2):
                low_sheet.cell(row=row, column=1).value = filename
                low_sheet.cell(row=row, column=2).value = author_name
                low_sheet.cell(row=row, column=3).value = word_count
                low_sheet.cell(row=row, column=4).value = min_words - word_count
        
        # 创建统计信息工作表
        stats_sheet = wb.create_sheet(title="统计信息")
        
        # 基本统计信息
        stats_data = [
            ["总文件数", len(word_counts)],
            ["平均字数", round(sum(x[1] for x in word_counts) / len(word_counts) if word_counts else 0, 1)],
            ["最大字数", max(x[1] for x in word_counts) if word_counts else 0],
            ["最小字数", min(x[1] for x in word_counts) if word_counts else 0],
            ["字数标准", min_words],
            ["字数不足文件数", len(low_wordcount_files)],
            ["字数合格文件数", len(word_counts) - len(low_wordcount_files)],
            ["合格率", f"{(1 - len(low_wordcount_files) / len(word_counts)) * 100:.1f}%" if word_counts else "0%"]
        ]
        
        for row, (label, value) in enumerate(stats_data, 1):
            stats_sheet.cell(row=row, column=1).value = label
            stats_sheet.cell(row=row, column=2).value = value
            stats_sheet.cell(row=row, column=1).font = Font(bold=True)
        
        # 自动调整列宽
        try:
            for sheet in wb.worksheets:
                # 创建列宽缓存以避免重复计算
                column_widths = {}
                
                for col in range(1, sheet.max_column + 1):
                    max_length = 0
                    column = get_column_letter(col)
                    column_widths[column] = max_length
                    
                    # 更高效的列宽计算
                    for cell in sheet[column]:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                                    column_widths[column] = max_length
                        except (TypeError, ValueError):
                            continue
                    
                    # 设置列宽，考虑中文字符宽度
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column].width = adjusted_width
        except Exception as e:
            print(f"! 调整Excel列宽时出错: {str(e)}")
        
        # 保存工作簿，增加异常处理
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            # 检查文件是否可写
            if os.path.exists(excel_path) and not os.access(excel_path, os.W_OK):
                alt_path = os.path.join(
                    os.path.dirname(excel_path),
                    f"字数统计报告_{current_time}_新.xlsx"
                )
                print(f"! 原文件无法写入，将保存到: {os.path.basename(alt_path)}")
                excel_path = alt_path
            
            wb.save(excel_path)
        except PermissionError:
            # 处理权限错误
            alt_path = os.path.join(
                os.path.dirname(excel_path),
                f"字数统计报告_{current_time}_新.xlsx"
            )
            print(f"! 保存Excel文件时出现权限错误，尝试保存到: {os.path.basename(alt_path)}")
            try:
                wb.save(alt_path)
                excel_path = alt_path
            except Exception as e:
                print(f"× Excel报告保存失败: {str(e)}")
                raise
        except Exception as e:
            print(f"× Excel报告保存失败: {str(e)}")
            raise
        
        return excel_path